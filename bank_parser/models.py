# Copyright 2024-2026 Koushik Mondal (github.com/raptar231)
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

from datetime import date
from decimal import Decimal
from typing import TYPE_CHECKING, Any

from pydantic import BaseModel, Field, field_validator

if TYPE_CHECKING:
    import pandas as pd  # type: ignore[import-untyped]


class Transaction(BaseModel):
    date: date
    description: str
    debit: Decimal | None = Field(default=None, ge=0)
    credit: Decimal | None = Field(default=None, ge=0)
    balance: Decimal = Field(ge=0)
    ref_no: str | None = None
    category: str = Field(pattern="^(debit|credit)$")

    @field_validator("debit", "credit", "balance", mode="before")
    @classmethod
    def parse_decimal(cls, v: str | Decimal | None) -> Decimal | None:
        if v is None or v == "":
            return None
        if isinstance(v, str):
            v = v.replace(",", "").strip()
            if v == "":
                return None
        return Decimal(str(v))

    @field_validator("category", mode="before")
    @classmethod
    def infer_category(cls, v: str | None, info: Any) -> str:
        if v:
            return v
        debit = info.data.get("debit")
        credit = info.data.get("credit")
        if debit is not None:
            return "debit"
        if credit is not None:
            return "credit"
        return "debit"

    def to_dict(self) -> dict:
        return {
            "date": self.date.isoformat(),
            "description": self.description,
            "debit": float(self.debit) if self.debit else "",
            "credit": float(self.credit) if self.credit else "",
            "balance": float(self.balance),
            "ref_no": self.ref_no or "",
            "category": self.category,
        }


class BalanceValidation(BaseModel):
    ok: bool = False
    status: str = "failed"
    reason: str | None = None
    expected_closing: Decimal | None = None
    calculated_closing: Decimal | None = None
    difference: Decimal | None = None
    total_debits: Decimal | None = None
    total_credits: Decimal | None = None


class Statement(BaseModel):
    bank: str
    account_number: str | None = None
    account_type: str | None = None
    statement_period_start: date | None = None
    statement_period_end: date | None = None
    opening_balance: Decimal | None = None
    closing_balance: Decimal | None = None
    transactions: list[Transaction] = []
    validation: BalanceValidation | None = None

    def validate_balances(self) -> BalanceValidation:
        """Reconcile opening + credits - debits against the statement's closing
        balance. Shared by every bank parser so a failed reconciliation is
        reported the same way (expected vs calculated vs difference)."""
        if self.opening_balance is None or self.closing_balance is None:
            result = BalanceValidation(
                ok=False,
                status="skipped",
                reason="opening/closing balance not extracted from statement",
            )
        elif not self.transactions:
            result = BalanceValidation(
                ok=False,
                status="skipped",
                reason="no transactions parsed from statement",
            )
        else:
            total_debits = Decimal("0")
            total_credits = Decimal("0")
            for txn in self.transactions:
                if txn.debit is not None:
                    total_debits += txn.debit
                if txn.credit is not None:
                    total_credits += txn.credit
            calculated = self.opening_balance + total_credits - total_debits
            difference = calculated - self.closing_balance
            result = BalanceValidation(
                ok=difference == 0,
                status="ok" if difference == 0 else "failed",
                expected_closing=self.closing_balance,
                calculated_closing=calculated,
                difference=difference,
                total_debits=total_debits,
                total_credits=total_credits,
            )
        self.validation = result
        return result

    def to_dataframe(self) -> "pd.DataFrame":  # type: ignore[name-defined]
        import pandas as pd  # type: ignore[import-untyped]

        return pd.DataFrame([t.to_dict() for t in self.transactions])

    def to_csv(self, path: str | None = None, **kwargs: Any) -> str | None:
        return self.to_dataframe().to_csv(path, **kwargs)  # type: ignore[no-any-return]

    def to_json(self, path: str | None = None, **kwargs: Any) -> str | None:
        return self.to_dataframe().to_json(path, **kwargs)  # type: ignore[no-any-return]

    def to_excel(self, path: str | None = None, **kwargs: Any) -> bytes | None:
        import openpyxl  # type: ignore[import-untyped]

        df = self.to_dataframe()
        wb = openpyxl.Workbook()

        # Transactions sheet
        ws_txn = wb.active
        ws_txn.title = "Transactions"
        for col_idx, col_name in enumerate(df.columns, 1):
            ws_txn.cell(row=1, column=col_idx, value=col_name)
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                ws_txn.cell(row=row_idx, column=col_idx, value=value)

        # Summary sheet
        ws_sum = wb.create_sheet("Summary")
        ws_sum.cell(row=1, column=1, value="Field")
        ws_sum.cell(row=1, column=2, value="Value")
        summary_data = [
            ("Bank", self.bank),
            ("Account Number", self.account_number or ""),
            ("Account Type", self.account_type or ""),
            (
                "Statement Period Start",
                str(self.statement_period_start) if self.statement_period_start else "",
            ),
            (
                "Statement Period End",
                str(self.statement_period_end) if self.statement_period_end else "",
            ),
            ("Opening Balance", float(self.opening_balance) if self.opening_balance else ""),
            ("Closing Balance", float(self.closing_balance) if self.closing_balance else ""),
            ("Total Transactions", len(self.transactions)),
            (
                "Total Debits",
                (
                    float(self.validation.total_debits)
                    if self.validation and self.validation.total_debits
                    else ""
                ),
            ),
            (
                "Total Credits",
                (
                    float(self.validation.total_credits)
                    if self.validation and self.validation.total_credits
                    else ""
                ),
            ),
            ("Validation Status", self.validation.status if self.validation else "N/A"),
        ]
        for row_idx, (field, value) in enumerate(summary_data, 2):
            ws_sum.cell(row=row_idx, column=1, value=field)
            ws_sum.cell(row=row_idx, column=2, value=value)

        if path:
            wb.save(path)
            return None
        else:
            from io import BytesIO

            bio = BytesIO()
            wb.save(bio)
            return bio.getvalue()

    def to_yaml(self, path: str | None = None, **kwargs: Any) -> str | None:
        import yaml  # type: ignore[import-untyped]

        data = {
            "bank": self.bank,
            "account_number": self.account_number,
            "account_type": self.account_type,
            "statement_period_start": (
                str(self.statement_period_start) if self.statement_period_start else None
            ),
            "statement_period_end": (
                str(self.statement_period_end) if self.statement_period_end else None
            ),
            "opening_balance": float(self.opening_balance) if self.opening_balance else None,
            "closing_balance": float(self.closing_balance) if self.closing_balance else None,
            "transactions": [t.to_dict() for t in self.transactions],
            "validation": (
                {
                    "ok": self.validation.ok if self.validation else False,
                    "status": self.validation.status if self.validation else "N/A",
                    "reason": self.validation.reason if self.validation else None,
                    "expected_closing": (
                        float(self.validation.expected_closing)
                        if self.validation and self.validation.expected_closing
                        else None
                    ),
                    "calculated_closing": (
                        float(self.validation.calculated_closing)
                        if self.validation and self.validation.calculated_closing
                        else None
                    ),
                    "difference": (
                        float(self.validation.difference)
                        if self.validation and self.validation.difference
                        else None
                    ),
                    "total_debits": (
                        float(self.validation.total_debits)
                        if self.validation and self.validation.total_debits
                        else None
                    ),
                    "total_credits": (
                        float(self.validation.total_credits)
                        if self.validation and self.validation.total_credits
                        else None
                    ),
                }
                if self.validation
                else None
            ),
        }

        yaml_str: str = yaml.dump(data, default_flow_style=False, sort_keys=False)

        if path:
            with open(path, "w") as f:
                f.write(yaml_str)
            return None
        else:
            return yaml_str


class GSTR2AEntry(BaseModel):
    gstin: str
    invoice_date: date
    invoice_number: str
    invoice_value: Decimal
    place_of_supply: str
    reverse_charge: str = "N"
    invoice_type: str = "B2B"
    rate: Decimal
    taxable_value: Decimal
    igst: Decimal = Decimal("0")
    cgst: Decimal = Decimal("0")
    sgst: Decimal = Decimal("0")
    cess: Decimal = Decimal("0")

    def to_dict(self) -> dict:
        return {
            "GSTIN": self.gstin,
            "Invoice Date": self.invoice_date.strftime("%d-%m-%Y"),
            "Invoice Number": self.invoice_number,
            "Invoice Value": float(self.invoice_value),
            "Place of Supply": self.place_of_supply,
            "Reverse Charge": self.reverse_charge,
            "Invoice Type": self.invoice_type,
            "Rate": float(self.rate),
            "Taxable Value": float(self.taxable_value),
            "IGST": float(self.igst),
            "CGST": float(self.cgst),
            "SGST": float(self.sgst),
            "CESS": float(self.cess),
        }
